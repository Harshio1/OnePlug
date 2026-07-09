import os
import shutil
import logging
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, BackgroundTasks, status
from sqlalchemy.orm import Session
from ..db import get_db
from ..config import settings
from .. import schemas, models
from ..services import db_service
from app.services.whisper import WhisperService
from app.services.analysis_service import AnalysisService
from app.services.translation_service import TranslationService
from app.services.gemini_service import GeminiService
from .auth import get_current_user
from app.integrations.myoperator import sync_recent_calls

router = APIRouter(prefix="/transcribe", tags=["transcription"])
logger = logging.getLogger(__name__)

# Instantiate services (singleton-like for the router)
whisper_service = WhisperService()
analysis_service = AnalysisService()
translation_service = TranslationService()
gemini_service = GeminiService()

PRIVILEGED_ROLES = {"admin", "manager"}

def can_access_audio(audio_file: models.AudioFile, user: models.User) -> bool:
    return user.role in PRIVILEGED_ROLES or audio_file.uploaded_by_id == user.id

def get_authorized_audio(db: Session, file_id: str, user: models.User) -> models.AudioFile:
    db_file = db_service.get_audio_file(db, file_id)
    if not db_file:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audio transcript not found.")
    if not can_access_audio(db_file, user):
        # Do not disclose whether another employee's call exists.
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Audio transcript not found.")
    return db_file

def process_transcription_task(file_id: str, language_hint: str = None, prompt: str = None):
    """
    Background worker that runs the OpenAI Whisper transcription
    and stores results in the PostgreSQL database.
    """
    # Create its own DB session for safety in background threads
    from ..db import SessionLocal
    db = SessionLocal()
    
    try:
        logger.info(f"Background task: Starting transcription for audio ID: {file_id}")
        db_file = db_service.get_audio_file(db, file_id)
        if not db_file:
            logger.error(f"Background task failed: Audio file {file_id} not found in database.")
            return

        db_service.update_audio_file_status(db, file_id, "processing")
        # Close the DB session during long CPU transcription to prevent holding SQLite locks
        db.close()
        
        # Execute Whisper API transcription (transcribes raw audio)
        result = whisper_service.transcribe_audio(
            file_path=db_file.file_path,
            language_hint=language_hint,
            prompt=prompt
        )
        
        # Reopen a new DB session to store the transcription results
        db = SessionLocal()
        
        # ── Gemini AI Processing ──────────────────────────────────────
        logger.info(f"Background task: Running Gemini AI for audio ID: {file_id}")
        raw_segments = result.get("segments", [])
        analysis_result = None
        if gemini_service.is_available:
            try:
                gemini_result = gemini_service.process_transcript(raw_segments)
                
                # Extract structured data
                english_transcript = gemini_result.get("transcript", result.get("text", ""))
                cleaned_segments = gemini_result.get("segments", raw_segments)
                
                # Map Gemini issue format to expected DB format
                issues = []
                for issue in gemini_result.get("issues", []):
                    issues.append({"issue_type": issue.get("type", "Unknown"), "severity": issue.get("severity", "Medium")})
                    
                issue_detected = len(issues) > 0
                    
                analysis_result = {
                    "analysed": True,
                    "summary": gemini_result.get("summary", ""),
                    "main_concern": gemini_result.get("main_concern", "See summary"),
                    "outcome": gemini_result.get("outcome", "Pending"),
                    "action_needed": gemini_result.get("action_needed", "Follow up as per summary."),
                    "what_happened": gemini_result.get("what_happened", "No breakdown available."),
                    "issue_detected": issue_detected,
                    "issue_type": issues[0]["issue_type"] if issue_detected else None,
                    "severity": issues[0]["severity"] if issue_detected else None,
                    "all_issues": issues,
                    "sentiment": gemini_result.get("sentiment", "Neutral"),
                    "sentiment_score": 0.0
                }
                
                # Overwrite the raw Whisper transcript and segments with the cleaned ones from Gemini
                result["text"] = english_transcript
                result["segments"] = cleaned_segments
                result["words_count"] = len(english_transcript.split())
            except Exception as ge:
                logger.error(f"Gemini processing failed, falling back to Whisper raw text: {ge}")

        if not analysis_result:
            logger.warning("Falling back to default empty analysis due to Gemini unavailability or error.")
            analysis_result = {
                "analysed": False,
                "summary": "AI Analysis fallback (Gemini API limit or quota reached).",
                "main_concern": "None",
                "outcome": "Pending",
                "action_needed": "Configure or check GEMINI_API_KEY.",
                "issue_detected": False,
                "issue_type": None,
                "severity": None,
                "all_issues": [],
                "sentiment": "Neutral",
                "sentiment_score": 0.0
            }

        logger.info(f"Background task: Gemini Processing complete for audio ID: {file_id}")

        # ── Save to DB ───────────────────────────────────────────────
        db_service.create_transcript(
            db=db,
            audio_file_id=file_id,
            text=result["text"],
            language=result["language"],
            words_count=result["words_count"],
            duration=result["duration"],
            segments=result["segments"],
            segment_count=result.get("segment_count", len(result["segments"])),
            analysis=analysis_result
        )
        logger.info(f"Background task: Complete for audio ID {file_id}.")
        
    except Exception as e:
        logger.error(f"Background task failed for audio ID {file_id}: {str(e)}")
        db_service.update_audio_file_status(
            db=db,
            file_id=file_id,
            status="failed",
            error_message=str(e)
        )
    finally:
        db.close()

@router.post("/upload", response_model=schemas.AudioFileResponse, status_code=status.HTTP_202_ACCEPTED)
def upload_audio(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    language_hint: Optional[str] = Form(None),
    prompt: Optional[str] = Form(None),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Upload an audio file (MP3/WAV) and queue it for Whisper transcription in a background thread.
    Returns the file metadata and processing status immediately.
    """
    # Verify file mime type or extension
    allowed_extensions = {".mp3", ".wav", ".m4a", ".ogg", ".flac"}
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported file format. Supported formats: {', '.join(allowed_extensions)}"
        )
        
    # Ensure upload directory exists
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    
    if file.content_type and not file.content_type.startswith("audio/"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded content must be an audio file.")

    # Save physical file to disk with a server-generated name and a bounded size.
    import uuid
    file_uuid = str(uuid.uuid4())
    secure_filename = f"{file_uuid}{file_ext}"
    dest_path = os.path.join(settings.UPLOAD_DIR, secure_filename)
    
    # Calculate file size as we write
    file_size = 0
    try:
        with open(dest_path, "wb") as buffer:
            while chunk := file.file.read(1024 * 1024):
                file_size += len(chunk)
                if file_size > settings.max_upload_size_bytes:
                    raise HTTPException(
                        status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                        detail=f"Audio exceeds the {settings.MAX_UPLOAD_SIZE_MB} MB upload limit."
                    )
                buffer.write(chunk)
    except Exception as e:
        if os.path.exists(dest_path):
            os.remove(dest_path)
        if isinstance(e, HTTPException):
            raise e
        logger.error(f"Failed to write uploaded file to disk: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to save uploaded file on backend server."
        )

    # Log file metadata to DB
    try:
        db_file = db_service.create_audio_file(
            db=db,
            filename=file.filename,
            file_path=dest_path,
            file_size=file_size,
            mime_type=file.content_type or f"audio/{file_ext[1:]}",
            user_id=current_user.id
        )
        
        # Schedule the transcription worker in the background
        background_tasks.add_task(
            process_transcription_task,
            file_id=db_file.id,
            language_hint=language_hint,
            prompt=prompt
        )
        
        return db_file
    except Exception as e:
        if os.path.exists(dest_path):
            os.remove(dest_path)
        logger.error(f"Failed to register audio file in database: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Database register failure. File rolled back."
        )

@router.get("/list", response_model=List[schemas.AudioFileListResponse])
def list_audios(
    skip: int = 0,
    limit: int = 2000,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Fetch a list of all audio files and their current transcription states.
    """
    limit = min(limit, 5000)
    owner_id = None if current_user.role in PRIVILEGED_ROLES else current_user.id
    return db_service.get_audio_files(db, skip=skip, limit=limit, uploaded_by_id=owner_id)

@router.get("/file/{file_id}", response_model=schemas.AudioFileResponse)
def get_audio_detail(
    file_id: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Retrieve single audio metadata and associated transcript contents.
    """
    return get_authorized_audio(db, file_id, current_user)

@router.delete("/delete/{file_id}", status_code=status.HTTP_200_OK)
def delete_audio(
    file_id: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Delete an audio upload, its local file representation, and associated transcripts.
    """
    db_file = get_authorized_audio(db, file_id, current_user)
        
    # Delete physically
    if os.path.exists(db_file.file_path):
        try:
            os.remove(db_file.file_path)
        except Exception as e:
            logger.error(f"Failed to remove physical file {db_file.file_path}: {e}")
            
    # Delete database representation (cascades to transcripts table)
    db.delete(db_file)
    db.commit()
    
    return {"status": "success", "message": f"Successfully deleted transcript and files for ID: {file_id}"}

from fastapi.responses import FileResponse

@router.get("/audio/{file_id}")
def stream_audio(
    file_id: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Streams the physical audio recording file for browser playback.
    """
    db_file = get_authorized_audio(db, file_id, current_user)
    if not os.path.exists(db_file.file_path):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Physical audio recording not found on disk."
        )
    return FileResponse(
        path=db_file.file_path,
        media_type=db_file.mime_type,
        filename=db_file.filename
    )

def run_sync_in_background(transcribe_task_fn):
    """Wrapper to run sync in background with its own DB session."""
    from ..db import SessionLocal
    db = SessionLocal()
    bg_tasks = BackgroundTasks()
    try:
        logger.info("Starting background MyOperator sync...")
        sync_recent_calls(db, bg_tasks, transcribe_task_fn)
        # Execute the queued transcription tasks one by one
        for task in bg_tasks.tasks:
            try:
                task.func(*task.args, **task.kwargs)
            except Exception as te:
                logger.error(f"Error executing queued transcription task: {te}")
    except Exception as e:
        logger.error(f"Background MyOperator sync failed: {e}")
    finally:
        db.close()

@router.post("/sync-myoperator", status_code=status.HTTP_202_ACCEPTED)
def trigger_myoperator_sync(
    background_tasks: BackgroundTasks,
    current_user: models.User = Depends(get_current_user),
):
    """
    Trigger manual / batch sync of recent call logs from MyOperator.
    Queues recording downloads and Whisper transcription.
    """
    # Verify user is admin or manager
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Unauthorized role for sync action.")
        
    background_tasks.add_task(run_sync_in_background, process_transcription_task)
    return {"status": "success", "message": "MyOperator sync started in the background."}


@router.post("/upload-customers")
async def upload_customers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Upload Excel file with customer data and store in customers table."""
    import openpyxl
    import io

    if current_user.role not in PRIVILEGED_ROLES:
        raise HTTPException(status_code=403, detail="Only admins or managers can upload customer data.")

    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
        raise HTTPException(status_code=400, detail="Only Excel or CSV files are supported.")

    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active

    # Find header row (row 2 based on your Excel structure)
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[2]]

    def col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    def normalize_phone(phone):
        if not phone:
            return None
        phone = str(phone).strip().replace(" ", "").replace("-", "")
        if phone.startswith("+91"):
            phone = phone[3:]
        elif phone.startswith("91") and len(phone) == 12:
            phone = phone[2:]
        elif phone.startswith("0") and len(phone) == 11:
            phone = phone[1:]
        return "+91" + phone if len(phone) == 10 else None

    imported = 0
    skipped = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        mobile_raw = row[col("Mobile Number")] if col("Mobile Number") is not None else None
        mobile = normalize_phone(mobile_raw)
        if not mobile:
            skipped += 1
            continue

        start_date = str(row[col("Start Date")]).strip() if col("Start Date") is not None and row[col("Start Date")] else None
        customer_name = str(row[col("Customer Name")]).strip() if col("Customer Name") is not None and row[col("Customer Name")] else None

        # Check if same customer + same date already exists
        existing = db.query(models.Customer).filter(
            models.Customer.mobile_number == mobile,
            models.Customer.start_date == start_date
        ).first()

        data = {
            "mobile_number": mobile,
            "customer_name": customer_name,
            "register_no": str(row[col("Register No")]).strip() if col("Register No") is not None and row[col("Register No")] else None,
            "station_name": str(row[col("Station Name")]).strip() if col("Station Name") is not None and row[col("Station Name")] else None,
            "location": str(row[col("Location")]).strip() if col("Location") is not None and row[col("Location")] else None,
            "start_date": start_date,
            "start_soc": str(row[col("Start SOC (%)")]).strip() if col("Start SOC (%)") is not None and row[col("Start SOC (%)")] else None,
            "end_soc": str(row[col("End SOC (%)")]).strip() if col("End SOC (%)") is not None and row[col("End SOC (%)")] else None,
            "total_units": str(row[col("Total Units")]).strip() if col("Total Units") is not None and row[col("Total Units")] else None,
            "vehicle_make": str(row[col("Vehicle Make")]).strip() if col("Vehicle Make") is not None and row[col("Vehicle Make")] else None,
            "vehicle_modal": str(row[col("Vehicle Modal")]).strip() if col("Vehicle Modal") is not None and row[col("Vehicle Modal")] else None,
            "charger_ownership": str(row[col("Charger Ownership")]).strip() if col("Charger Ownership") is not None and row[col("Charger Ownership")] else None,
            "rating_feedback": str(row[col("Rating feedback")]).strip() if col("Rating feedback") is not None and row[col("Rating feedback")] else None,
            "last_transaction_date": str(row[col("Last Transaction Date")]).strip() if col("Last Transaction Date") is not None and row[col("Last Transaction Date")] else None,
            "number_of_rating_stars": str(row[col("Number of Rating Stars")]).strip() if col("Number of Rating Stars") is not None and row[col("Number of Rating Stars")] else None,
            "rating_comments": str(row[col("Rating Comments")]).strip() if col("Rating Comments") is not None and row[col("Rating Comments")] else None,
            "agent_name": str(row[col("Call Center Agent Name")]).strip() if col("Call Center Agent Name") is not None and row[col("Call Center Agent Name")] else None,
        }

        if existing:
            for k, v in data.items():
                if v is not None:
                    setattr(existing, k, v)
            db.commit()
        else:
            new_customer = models.Customer(**data)
            db.add(new_customer)
            db.commit()

        imported += 1

    return {"success": True, "imported": imported, "skipped": skipped}
